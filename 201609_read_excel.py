import json
import pprint

from openpyxl import load_workbook
wb = load_workbook('video_sync.xlsx')
#print wb2.get_sheet_names()
ws = wb.active

title          = ws["B1"].value.encode('utf-8')
slide_url      = ws["B2"].value.encode('utf-8')
video_url      = ws["B3"].value.encode('utf-8')
total_duration = ws["B4"].value * 60 + ws["C4"].value

cell_idx = 7
json_data = []
while ws["A"+str(cell_idx)].value is not None:
    #print ws["B"+str(cell_idx)].value
    min = ws["B"+str(cell_idx)].value or 0
    sec = ws["C"+str(cell_idx)].value or 0
    json_data.append({"url": slide_url+"#"+str(ws["A"+str(cell_idx)].value), 
                        "time" : min*60 + sec})
    cell_idx += 1

result_str = """
{{
    "title": "{title}",
    "chapters": [
    {{
        "title": "{title}",
        "duration": {total_duration},
        "video": {{
            "url": "{video_url}"
        }},
        "slides": {result_json}}}
    ]
}}""".format(
    title = title,
    total_duration = total_duration,
    video_url = video_url,
    result_json = json.dumps(json_data, indent=8, sort_keys=True)
)

print result_str
#print json.dumps(json.loads(result_str), indent=4, sort_keys=False)
